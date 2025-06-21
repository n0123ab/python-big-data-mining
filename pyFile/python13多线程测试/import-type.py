import openpyxl
import random
import os

def generate_unique_19_digit(existing_numbers):
    while True:
        num = str(random.randint(10**18, 10**19 - 1))
        if num not in existing_numbers:
            return num

def main():
    file_path = 'import.xlsx'
    if not os.path.exists(file_path):
        print(f"文件 {file_path} 不存在。")
        return

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # 收集已有的19位数字
    existing_numbers = set()
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        cell_value = str(row[0].value).strip() if row[0].value is not None else ''
        if cell_value.isdigit() and len(cell_value) == 19:
            existing_numbers.add(cell_value)

    # 填充缺失的19位数字
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        cell = row[0]
        cell_value = str(cell.value).strip() if cell.value is not None else ''
        if not (cell_value.isdigit() and len(cell_value) == 19):
            new_num = generate_unique_19_digit(existing_numbers)
            cell.value = new_num
            existing_numbers.add(new_num)

    wb.save(file_path)
    print("处理完成，已保存。")

if __name__ == "__main__":
    main()